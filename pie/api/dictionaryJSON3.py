

import openpyxl






def openfile(wb):
    while True:
        try:
            workbook = openpyxl.load_workbook(wb)
        except FileNotFoundError:
            return None
        else:
            sheet = workbook.get_sheet_by_name('Sheet1')
            return sheet

def findyearrange(sheet):
    yearrange = []
    for i in range(2,sheet.max_row+1):
        thiscell = sheet.cell(row=i,column=1).value
        if not thiscell in yearrange:
            if not thiscell == None:
                yearrange.append(thiscell)
    return yearrange


def findcatrange(sheet):
    catrange = []
    for i in range(2,sheet.max_row+1):
        thiscell = sheet.cell(row=i,column=2).value
        if not thiscell in catrange:
            if not thiscell == None:
                catrange.append(thiscell)
    
    return catrange





def catsum(category,sheet):
    total = 0
    
    for i in range(2,sheet.max_row+1):
        thiscell = sheet.cell(row=i,column=2).value
        if thiscell == category:
            total += sheet.cell(row=i,column=3).value
    
    return total

###produces list of totals in corresponding order to category list
def cattotals(sheet,cats):
    totals = []
    for i in cats:
        totals.append(catsum(i,sheet))
    #print(totals)
    return totals



##produce sorted list of categories

def sortcats(cats,categorytotals):
    
    return [i[0] for i in (list(sorted(zip(cats,categorytotals),key = lambda x:x[1],reverse=True)))]



def add_colors(colors,cats):
    colors = colors*int((len(cats)/len(colors)+1))
    colors = colors[:len(cats)]
    d = {k:v for k,v in zip(cats,colors)}
    return d




def getHeaders(sheet):
    headers = []
    for i in range(1,sheet.max_column):
        headers.append(sheet.cell(row=1,column=i).value)
    return headers

#headers = getHeaders()
headers = ["year","name","count","color"]
#print("The headers are: ",headers)

#####the next part needs to know year range, sortedcats and headers - also the color key

def main(year,headers,order,color_key,sheet):
    def rownumbers(year):
        rownums = []
        catnames = []
        for i in range(1,sheet.max_row+1):
            if sheet.cell(row=i,column=1).value==year:
               rownums.append([(headers[1],sheet.cell(row=i,column=2).value),(headers[2],sheet.cell(row=i,column=3).value),(headers[3],color_key[sheet.cell(row=i,column=2).value])])##list of tuples each represents a row for the given year
                ###first tuple above is category - you can use that as the key to get the corresponding color
               catnames.append(sheet.cell(row=i,column=2).value)
        return rownums, catnames
    ####get all values, in order matching category order for that year.

    rows,cats = rownumbers(year)


    def sorted_data_by_year():
        ordered_rows = []
        for o in order:
            if o in cats:
                ordered_rows.append(rows[cats.index(o)])
            else:
                ordered_rows.append([(headers[1],o),(headers[2],0),(headers[3],color_key[o])])###this is where blank values are added
        #print("Ordered Rows",ordered_rows)
        return ordered_rows

    sorted_data = sorted_data_by_year()
    #print(year, sorted_data)
    return year,sorted_data

colors = ["rgb(130,34,171)","rgb(171,34,100)",
          "rgb(34, 141, 171)","rgb(57,112,241)","rgb(55, 191, 34)",
          "rgb(65,171,22)","rgb(57, 34, 171)","rgb(241,174,57)",
          "rgb(171,100,34)","rgb(171,34,34)","rgb(241,57,57)"]




def make_dict(year,datumyear):
    megadict=[]


    for data in datumyear:
        d = {k:v for k,v in data}

        megadict.append(d)

    proto = [{"year":year,"data":megadict}]
    
    return proto[0]

def JSONify(wb):
    

    sheet = openfile(wb)
    if sheet:
    
        years = findyearrange(sheet)
        years.sort()
        cats = findcatrange(sheet)
        
        
        color_key = add_colors(colors,cats)
        categorytotals = cattotals(sheet,cats)
        sortedcats = sortcats(cats,categorytotals)

        evenmoremegadict = []
        for y in years:
            year,datumyear = main(y,headers,sortedcats,color_key,sheet)
            evenmoremegadict.append(make_dict(year,datumyear))
        
        return evenmoremegadict
    else:
        return None

    

#JSONifydata = JSONify('missingpiecessorted2.xlsx')
#print(JSONifydata)
